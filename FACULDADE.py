import tkinter as tk
from tkinter import ttk
from tkinter import messagebox, filedialog
from datetime import datetime
import mysql.connector
import pandas as pd
import openpyxl
from openpyxl.styles import Font

# Função para conectar ao banco de dados
def conectar_db():
    try:
        return mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="faculdade_1"
        )
    except mysql.connector.Error as err:
        messagebox.showerror("Erro de Conexão", f"Erro ao conectar ao banco de dados: {err}")
        return None

# Função para limpar o frame atual
def clear_frame():
    for widget in root.winfo_children():
        widget.destroy()

# Função para voltar ao menu principal
def voltar_menu():
    clear_frame()
    menu_frame = ttk.Frame(root, padding="20", style="Custom.TFrame")
    menu_frame.pack(fill="both", expand=True)

    ttk.Label(menu_frame, text="Menu Principal", font=("Helvetica", 24, "bold"), style="Custom.TLabel").pack(pady=20)

    btn_cobranca = ttk.Button(menu_frame, text="Cobrança", style="Custom.TButton", command=abrir_cobranca)
    btn_cobranca.pack(pady=10, fill='x')

    btn_pix = ttk.Button(menu_frame, text="PIX", style="Custom.TButton", command=abrir_pix)
    btn_pix.pack(pady=10, fill='x')

# Função para abrir o menu de cobrança
import mysql.connector
from openpyxl import load_workbook  # Adicionar a importação aqui
from datetime import datetime

def abrir_cobranca():
    # Função para conectar ao banco de dados
    def conectar_db():
        try:
            return mysql.connector.connect(
                host="localhost",
                user="root",
                password="",
                database="faculdade_1"
            )
        except mysql.connector.Error as err:
            print(f"Erro ao conectar ao banco de dados: {err}")
            return None

    # Função para coletar e processar os dados
    def coletar_dados():
        conn = conectar_db()
        if conn is None:
            return

        try:
            cursor = conn.cursor()

            # Consulta para pegar a quantidade de boletos por banco e mês
            cursor.execute("""
                SELECT REF, COD_BANCO, SUM(QTDE) AS QTDE
                FROM marketshare_bancos
                GROUP BY REF, COD_BANCO
                ORDER BY REF, COD_BANCO
            """)
            resultados = cursor.fetchall()

            # Estruturar dados em um dicionário com REF como chave principal
            data_dict = {}
            for ref, cod_banco, qtde in resultados:
                if ref not in data_dict:
                    data_dict[ref] = {}
                data_dict[ref][cod_banco] = qtde

            return data_dict

        except Exception as e:
            print(f"Erro ao coletar os dados: {e}")
        finally:
            cursor.close()
            conn.close()

    # Função para substituir os dados no Excel existente
    def substituir_dados_no_excel(data_dict):
        excel_file = "C:\\Users\\Leonardo\\Desktop\\relatorio_boletos_bancos.xlsx"

        # Carregar o arquivo Excel existente
        wb = load_workbook(excel_file)
        ws = wb.active
        ws.title = "Relatório de Boletos"

        # Cabeçalhos - converter as chaves para "Mês/Ano"
        headers = ["Banco \\ Mês"] + [ref.strftime("%m/%Y") for ref in sorted(data_dict.keys())]
        
        # Inserir o cabeçalho na primeira linha
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)

        # Linhas iniciais para Mercado Total (sem Bradesco) e Demais Bancos
        mercado_total = ["Mercado Total (sem Bradesco)"]
        demais_bancos = ["Demais Bancos"]

        for ref in data_dict.keys():
            # Quantidade total de boletos para todos os bancos, exceto o Bradesco
            bradesco_qtde = data_dict[ref].get(237, 0)
            total_ref = sum(data_dict[ref].values())
            demais_qtde = total_ref - bradesco_qtde

            mercado_total.append(demais_qtde)
            demais_bancos.append(demais_qtde)

        # Substituir as linhas de Mercado Total (sem Bradesco) e Demais Bancos
        for i, linha in enumerate([mercado_total, demais_bancos], start=2):  # Começar a partir da linha 2
            for j, valor in enumerate(linha, start=1):
                ws.cell(row=i, column=j, value=valor)

        # Obter todos os códigos de banco, incluindo o Bradesco (237)
        bancos = sorted({cod_banco for ref_data in data_dict.values() for cod_banco in ref_data})

        # Substituir os dados de cada banco na tabela detalhada, incluindo o Bradesco
        row_num = 4  # Começar na linha 4 para os dados
        for cod_banco in bancos:
            linha = [cod_banco]
            for ref in data_dict.keys():
                qtde = data_dict[ref].get(cod_banco, 0)
                linha.append(qtde)
            
            # Substituir a linha na planilha
            for col_num, valor in enumerate(linha, start=1):
                ws.cell(row=row_num, column=col_num, value=valor)
            
            row_num += 1

        # Salvar as alterações no arquivo Excel
        wb.save(excel_file)
        messagebox.showinfo("Sucesso", "Dados substituídos com sucesso no arquivo Excel.")

    # Executar as funções
    data = coletar_dados()
    if data:
        substituir_dados_no_excel(data)


# Função para abrir o menu Pix
def abrir_pix():
    clear_frame()
    frame_pix = ttk.Frame(root, padding="20", style="Custom.TFrame")
    frame_pix.pack(fill="both", expand=True)

    ttk.Label(frame_pix, text="Opções de PIX", font=("Helvetica", 20, "bold"), style="Custom.TLabel").pack(pady=20)

    btn_isentar = ttk.Button(frame_pix, text="Isentar Novas Contas", style="Custom.TButton", command=abrir_isencao)
    btn_isentar.pack(pady=10, fill='x')

    btn_verificar = ttk.Button(frame_pix, text="Verificar Data de Vencimento", style="Custom.TButton", command=verificar_data_vencimento)
    btn_verificar.pack(pady=10, fill='x')

# Função para abrir o menu de isenção
def abrir_isencao():
    clear_frame()
    frame_isencao = ttk.Frame(root, padding="20", style="Custom.TFrame")
    frame_isencao.pack(fill="both", expand=True)

    ttk.Label(frame_isencao, text="Isentar Novas Contas", font=("Helvetica", 20, "bold"), style="Custom.TLabel").pack(pady=20)

    # Opções de inserção
    btn_inserir_manual = ttk.Button(frame_isencao, text="Inserir Manualmente", style="Custom.TButton", command=inserir_manual)
    btn_inserir_manual.pack(pady=10)

    btn_selecionar_arquivo = ttk.Button(frame_isencao, text="Selecionar Arquivo", style="Custom.TButton", command=selecionar_arquivo)
    btn_selecionar_arquivo.pack(pady=10)

    # Botão Voltar
    btn_voltar = ttk.Button(frame_isencao, text="Voltar ao Menu", style="Custom.TButton", command=voltar_menu)
    btn_voltar.pack(pady=10)

# Função para inserir manualmente
def inserir_manual():
    clear_frame()
    frame_isencao = ttk.Frame(root, padding="20", style="Custom.TFrame")
    frame_isencao.pack(fill="both", expand=True)

    ttk.Label(frame_isencao, text="Isentar Novas Contas - Inserção Manual", font=("Helvetica", 20, "bold"), style="Custom.TLabel").pack(pady=20)

    # Perguntas para isentar manualmente
    ttk.Label(frame_isencao, text="Agência:", style="Custom.TLabel").pack(pady=5)
    agencia_entry = ttk.Entry(frame_isencao, font=("Helvetica", 14))
    agencia_entry.pack(pady=5)

    ttk.Label(frame_isencao, text="Conta:", style="Custom.TLabel").pack(pady=5)
    conta_entry = ttk.Entry(frame_isencao, font=("Helvetica", 14))
    conta_entry.pack(pady=5)

    tarifas_list = []

    def adicionar_tarifa():
        tarifa = tarifa_entry.get()
        if tarifa:
            tarifas_list.append(tarifa)
            tarifa_listbox.insert(tk.END, tarifa)
            tarifa_entry.delete(0, tk.END)  # Limpa o campo de entrada da tarifa

    def remover_tarifa():
        try:
            selected_index = tarifa_listbox.curselection()[0]
            tarifa_listbox.delete(selected_index)
            tarifas_list.pop(selected_index)  # Remove a tarifa da lista
        except IndexError:
            messagebox.showwarning("Aviso", "Selecione uma tarifa para remover.")

    ttk.Label(frame_isencao, text="Tarifa:", style="Custom.TLabel").pack(pady=5)
    tarifa_entry = ttk.Entry(frame_isencao, font=("Helvetica", 14))
    tarifa_entry.pack(pady=5)

    tarifa_btn = ttk.Button(frame_isencao, text="Adicionar Tarifa", style="Custom.TButton", command=adicionar_tarifa)
    tarifa_btn.pack(pady=10)

    tarifa_listbox = tk.Listbox(frame_isencao, font=("Helvetica", 12), height=5)
    tarifa_listbox.pack(pady=5)

    # Botão para remover tarifa
    remover_tarifa_btn = ttk.Button(frame_isencao, text="Remover Tarifa", style="Custom.TButton", command=remover_tarifa)
    remover_tarifa_btn.pack(pady=10)

    ttk.Label(frame_isencao, text="Período (dias):", style="Custom.TLabel").pack(pady=5)
    periodo_entry = ttk.Entry(frame_isencao, font=("Helvetica", 14))
    periodo_entry.pack(pady=5)

    ttk.Label(frame_isencao, text="Motivo da Isenção:", style="Custom.TLabel").pack(pady=5)
    motivo_combobox = ttk.Combobox(frame_isencao, values=["ISENCAO VIA MPI", "ISENCAO MEI/EI", "OUTROS"], font=("Helvetica", 14))
    motivo_combobox.pack(pady=5)

    def submeter_isencao():
        agencia = agencia_entry.get()
        conta = conta_entry.get()
        periodo = periodo_entry.get()
        motivo = motivo_combobox.get()

        if not agencia or not conta or not periodo or not motivo:
            messagebox.showerror("Erro", "Todos os campos devem ser preenchidos.")
            return

        try:
            periodo = int(periodo)
            if periodo <= 0:
                raise ValueError("O período deve ser um número positivo.")
        except ValueError:
            messagebox.showerror("Erro", "Período deve ser um número válido e positivo.")
            return

        data_deferimento = datetime.now().date()
        db = conectar_db()
        if db is None:
            return  # Se a conexão falhar, não prosseguir

        cursor = db.cursor()
        for tarifa in tarifas_list:
            cursor.execute("""INSERT INTO tb_contas_isentar (AGENCIA, CONTA, PERIODO, FLEXIBILIZACAO, TARIFA, MOTIVO, DATA_DEFERIMENTO)
                              VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                           (agencia, conta, periodo, 100, tarifa, motivo, data_deferimento))

        db.commit()
        cursor.close()
        db.close()
        messagebox.showinfo("Sucesso", "Isenção cadastrada com sucesso.")

        # Limpar campos após cadastrar
        agencia_entry.delete(0, tk.END)
        conta_entry.delete(0, tk.END)
        periodo_entry.delete(0, tk.END)
        motivo_combobox.set('')
        tarifas_list.clear()
        tarifa_listbox.delete(0, tk.END)

    submit_btn = ttk.Button(frame_isencao, text="Cadastrar Isenção", style="Custom.TButton", command=submeter_isencao)
    submit_btn.pack(pady=20)

    # Botão Voltar
    btn_voltar = ttk.Button(frame_isencao, text="Voltar ao Menu", style="Custom.TButton", command=voltar_menu)
    btn_voltar.pack(pady=10)

# Função para selecionar arquivo
def selecionar_arquivo():
    clear_frame()
    frame_selecionar_arquivo = ttk.Frame(root, padding="20", style="Custom.TFrame")
    frame_selecionar_arquivo.pack(fill="both", expand=True)

    ttk.Label(frame_selecionar_arquivo, text="Selecionar Arquivo", font=("Helvetica", 20, "bold"), style="Custom.TLabel").pack(pady=20)

    tipo_tarifa_var = tk.StringVar()

    ttk.Label(frame_selecionar_arquivo, text="Selecione o Tipo da Tarifa:", style="Custom.TLabel").pack(pady=5)
    
    tipo_tarifa_pagamento = ttk.Radiobutton(frame_selecionar_arquivo, text="Pagamento", variable=tipo_tarifa_var, value="pagamento", style="Custom.TRadiobutton")
    tipo_tarifa_pagamento.pack(pady=5)
    
    tipo_tarifa_recebimento = ttk.Radiobutton(frame_selecionar_arquivo, text="Recebimento", variable=tipo_tarifa_var, value="recebimento", style="Custom.TRadiobutton")
    tipo_tarifa_recebimento.pack(pady=5)

    def carregar_excel():
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
        if not file_path:
            return

        try:
            df = pd.read_excel(file_path)

            if 'agencia' not in df.columns or 'conta' not in df.columns or 'periodo' not in df.columns or 'motivo' not in df.columns:
                messagebox.showerror("Erro", "O Excel deve conter as colunas: 'agencia', 'conta', 'periodo' e 'motivo'.")
                return

            db = conectar_db()
            if db is None:
                return  # Se a conexão falhar, não prosseguir

            cursor = db.cursor()
            tipo_tarifa = tipo_tarifa_var.get()
            tarifas = []

            if tipo_tarifa == "pagamento":
                tarifas = [3344, 3322, 3311, 3355]
            elif tipo_tarifa == "recebimento":
                tarifas = [4433, 4422, 4411, 4455]

            for index, row in df.iterrows():
                agencia = row['agencia']
                conta = row['conta']
                periodo = row['periodo']
                motivo = row['motivo']

                for tarifa in tarifas:
                    cursor.execute("""INSERT INTO tb_contas_isentar (AGENCIA, CONTA, PERIODO, FLEXIBILIZACAO, TARIFA, MOTIVO, DATA_DEFERIMENTO)
                                      VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                                   (agencia, conta, periodo, 100, tarifa, motivo, datetime.now().date()))

            db.commit()
            cursor.close()
            db.close()
            messagebox.showinfo("Sucesso", "Dados carregados e cadastrados com sucesso.")
            voltar_menu()  # Volta ao menu principal após o carregamento

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao processar o arquivo: {e}")

    btn_carregar_excel = ttk.Button(frame_selecionar_arquivo, text="Carregar Excel", style="Custom.TButton", command=carregar_excel)
    btn_carregar_excel.pack(pady=20)

    # Botão Voltar
    btn_voltar = ttk.Button(frame_selecionar_arquivo, text="Voltar ao Menu", style="Custom.TButton", command=voltar_menu)
    btn_voltar.pack(pady=10)

# Função para verificar a data de vencimento
def verificar_data_vencimento():
    clear_frame()
    frame_verificar = ttk.Frame(root, padding="20", style="Custom.TFrame")
    frame_verificar.pack(fill="both", expand=True)

    ttk.Label(frame_verificar, text="Verificar Data de Vencimento", font=("Helvetica", 20, "bold"), style="Custom.TLabel").pack(pady=20)

    ttk.Label(frame_verificar, text="Data de Vencimento (AAAA-MM-DD):", style="Custom.TLabel").pack(pady=5)
    vencimento_entry = ttk.Entry(frame_verificar, font=("Helvetica", 14))
    vencimento_entry.pack(pady=5)

    def verificar():
        vencimento_str = vencimento_entry.get()
        try:
            data_vencimento = datetime.strptime(vencimento_str, '%Y-%m-%d').date()
        except ValueError:
            messagebox.showerror("Erro", "Data deve estar no formato AAAA-MM-DD.")
            return

        db = conectar_db()
        if db is None:
            return  # Se a conexão falhar, não prosseguir

        cursor = db.cursor()
        cursor.execute("SELECT * FROM tb_contas_isentar WHERE DATA_DEFERIMENTO <= %s", (data_vencimento,))
        resultados = cursor.fetchall()

        if resultados:
            # Criar o arquivo Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Contas Isentas"

            # Cabeçalhos
            headers = ["Agência", "Conta", "Período", "Tarifa", "Motivo", "Data Deferimento"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)

            # Inserir dados
            for row_num, resultado in enumerate(resultados, 2):
                ws.cell(row=row_num, column=1, value=resultado[0])  # Agência
                ws.cell(row=row_num, column=2, value=resultado[1])  # Conta
                ws.cell(row=row_num, column=3, value=resultado[2])  # Período
                ws.cell(row=row_num, column=4, value=resultado[4])  # Tarifa
                ws.cell(row=row_num, column=5, value=resultado[5])  # Motivo
                ws.cell(row=row_num, column=6, value=resultado[6])  # Data Deferimento

            # Salvar arquivo
            try:
                file_name = f"contas_isentas_{data_vencimento}.xlsx"
                wb.save(file_name)
                messagebox.showinfo("Sucesso", f"Arquivo '{file_name}' criado com sucesso!")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao salvar o arquivo: {str(e)}")
        else:
            messagebox.showinfo("Resultados", "Nenhuma conta com vencimento até a data selecionada.")

        cursor.close()
        db.close()

    btn_verificar = ttk.Button(frame_verificar, text="Verificar", style="Custom.TButton", command=verificar)
    btn_verificar.pack(pady=20)

    # Botão Voltar
    btn_voltar = ttk.Button(frame_verificar, text="Voltar ao Menu", style="Custom.TButton", command=voltar_menu)
    btn_voltar.pack(pady=10)
# Configurações da janela principal
root = tk.Tk()
root.title("Sistema de Isenção de Contas")
root.geometry("600x600")

# Estilo da interface
style = ttk.Style()
style.configure("Custom.TFrame", background="#F0F0F0")
style.configure("Custom.TLabel", background="#F0F0F0", font=("Helvetica", 14))
style.configure("Custom.TButton", padding=10, font=("Helvetica", 14), background="#0071BC", foreground="black")
style.map("Custom.TButton", background=[("active", "#005FA3")])

# Inicia com o menu principal
voltar_menu()

# Executa a aplicação
root.mainloop()
